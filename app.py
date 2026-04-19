import os
import json
import base64
import uuid
import datetime
import anthropic
import openpyxl
import io
import sqlite3
from pathlib import Path
from functools import wraps

from flask import (
    Flask, request, jsonify, render_template,
    session, redirect, url_for, send_file, abort
)
from authlib.integrations.flask_client import OAuth

from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "dev-secret-change-me")

APP_URL = os.environ.get("APP_URL", "http://localhost:5000")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")
DATABASE_URL = os.environ.get("DATABASE_URL", "")

SCOPES = [
    "openid", "email", "profile",
    "https://www.googleapis.com/auth/drive",
]

oauth = OAuth(app)
google = oauth.register(
    name="google",
    client_id=GOOGLE_CLIENT_ID,
    client_secret=GOOGLE_CLIENT_SECRET,
    server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
    client_kwargs={"scope": " ".join(SCOPES), "access_type": "offline", "prompt": "consent"},
)

CATEGORY_MAP = {
    "food": "Food & Grocery", "tech": "Technology",
    "services": "Services", "health": "Health & Medical",
    "retail": "Retail", "other": "Other",
}

# ---------------------------------------------------------------------------
# Database — works with both PostgreSQL (via pg8000) and SQLite
# ---------------------------------------------------------------------------
def get_db():
    if DATABASE_URL and DATABASE_URL.startswith("postgres"):
        import pg8000.native
        url = DATABASE_URL
        if url.startswith("postgres://"):
            url = url.replace("postgres://", "postgresql://", 1)
        # Parse the URL manually for pg8000
        from urllib.parse import urlparse
        p = urlparse(url)
        return pg8000.native.Connection(
            host=p.hostname, port=p.port or 5432,
            database=p.path.lstrip("/"),
            user=p.username, password=p.password,
            ssl_context=True,
        ), "pg8000"
    else:
        db_path = Path("invoices.db")
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        return conn, "sqlite"


def init_db():
    conn, kind = get_db()
    if kind == "pg8000":
        conn.run("""CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY, email TEXT UNIQUE NOT NULL, name TEXT, picture TEXT,
            google_id TEXT UNIQUE, access_token TEXT, refresh_token TEXT,
            drive_folder_id TEXT, drive_folder_name TEXT,
            webhook_channel_id TEXT, webhook_expiry TEXT, created_at TEXT
        )""")
        conn.run("""CREATE TABLE IF NOT EXISTS invoices (
            id TEXT PRIMARY KEY, user_id TEXT NOT NULL, filename TEXT,
            drive_file_id TEXT, category TEXT, vendor TEXT, total TEXT,
            invoice_date TEXT, confidence INTEGER, summary TEXT,
            source TEXT DEFAULT 'manual', processed_at TEXT
        )""")
        conn.close()
    else:
        cur = conn.cursor()
        cur.execute("""CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY, email TEXT UNIQUE NOT NULL, name TEXT, picture TEXT,
            google_id TEXT UNIQUE, access_token TEXT, refresh_token TEXT,
            drive_folder_id TEXT, drive_folder_name TEXT,
            webhook_channel_id TEXT, webhook_expiry TEXT, created_at TEXT
        )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS invoices (
            id TEXT PRIMARY KEY, user_id TEXT NOT NULL, filename TEXT,
            drive_file_id TEXT, category TEXT, vendor TEXT, total TEXT,
            invoice_date TEXT, confidence INTEGER, summary TEXT,
            source TEXT DEFAULT 'manual', processed_at TEXT
        )""")
        conn.commit()
        conn.close()


def db_fetchone(query, params=()):
    conn, kind = get_db()
    try:
        if kind == "pg8000":
            rows = conn.run(query, *params)
            cols = [c["name"] for c in conn.columns]
            if rows:
                return dict(zip(cols, rows[0]))
            return None
        else:
            cur = conn.cursor()
            cur.execute(query, params)
            row = cur.fetchone()
            return dict(row) if row else None
    finally:
        conn.close()


def db_fetchall(query, params=()):
    conn, kind = get_db()
    try:
        if kind == "pg8000":
            rows = conn.run(query, *params)
            cols = [c["name"] for c in conn.columns]
            return [dict(zip(cols, r)) for r in rows]
        else:
            cur = conn.cursor()
            cur.execute(query, params)
            return [dict(r) for r in cur.fetchall()]
    finally:
        conn.close()


def db_execute(query, params=()):
    conn, kind = get_db()
    try:
        if kind == "pg8000":
            conn.run(query, *params)
        else:
            cur = conn.cursor()
            cur.execute(query, params)
            conn.commit()
    finally:
        conn.close()


def db_count(query, params=()):
    conn, kind = get_db()
    try:
        if kind == "pg8000":
            rows = conn.run(query, *params)
            return rows[0][0] if rows else 0
        else:
            cur = conn.cursor()
            cur.execute(query, params)
            return cur.fetchone()[0]
    finally:
        conn.close()


# Init DB on startup
with app.app_context():
    try:
        init_db()
    except Exception as e:
        app.logger.error(f"DB init error: {e}")


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated


def current_user():
    uid = session.get("user_id")
    if not uid:
        return None
    return db_fetchone("SELECT * FROM users WHERE id = ?", (uid,)) if not DATABASE_URL.startswith("postgres") \
        else db_fetchone("SELECT * FROM users WHERE id = :1", (uid,))


# ---------------------------------------------------------------------------
# Drive helpers
# ---------------------------------------------------------------------------
def get_drive_service(user):
    if not user or not user.get("access_token"):
        return None
    creds = Credentials(
        token=user["access_token"],
        refresh_token=user.get("refresh_token"),
        token_uri="https://oauth2.googleapis.com/token",
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    return build("drive", "v3", credentials=creds)


def register_webhook(user):
    service = get_drive_service(user)
    if not service or not user.get("drive_folder_id"):
        return False
    channel_id = str(uuid.uuid4())
    expiry = datetime.datetime.utcnow() + datetime.timedelta(days=7)
    body = {
        "id": channel_id, "type": "web_hook",
        "address": f"{APP_URL.rstrip('/')}/webhook/{user['id']}",
        "expiration": str(int(expiry.timestamp() * 1000)),
    }
    try:
        service.files().watch(fileId=user["drive_folder_id"], body=body).execute()
        db_execute(
            "UPDATE users SET webhook_channel_id=?, webhook_expiry=? WHERE id=?",
            (channel_id, expiry.isoformat(), user["id"])
        )
        return True
    except Exception as e:
        app.logger.error(f"Webhook error: {e}")
        return False


def run_claude(content_blocks):
    content_blocks.append({"type": "text", "text": """Analyze this invoice. Respond ONLY with valid JSON (no markdown):
{"category":"food|tech|services|health|retail|other","confidence":85,"vendor":"name or N/A","total":"amount+currency or N/A","date":"date or N/A","summary":"2-3 sentences describing the invoice and why you chose this category"}"""})
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    msg = client.messages.create(
        model="claude-sonnet-4-20250514", max_tokens=800,
        messages=[{"role": "user", "content": content_blocks}],
    )
    raw = "".join(b.text for b in msg.content if hasattr(b, "text"))
    return json.loads(raw.replace("```json", "").replace("```", "").strip())


def save_invoice(user_id, result, filename, drive_file_id=None, source="manual"):
    inv_id = str(uuid.uuid4())
    db_execute(
        "INSERT INTO invoices (id,user_id,filename,drive_file_id,category,vendor,total,invoice_date,confidence,summary,source,processed_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        (inv_id, user_id, filename, drive_file_id,
         result.get("category", "other"), result.get("vendor", ""),
         result.get("total", ""), result.get("date", ""),
         int(result.get("confidence", 0)), result.get("summary", ""),
         source, datetime.datetime.utcnow().isoformat())
    )
    return inv_id


# ---------------------------------------------------------------------------
# Routes — Auth
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    if session.get("user_id"):
        return redirect(url_for("dashboard"))
    return render_template("landing.html")


@app.route("/login")
def login_page():
    return render_template("landing.html")


@app.route("/auth/google")
def auth_google():
    redirect_uri = url_for("auth_callback", _external=True)
    return google.authorize_redirect(redirect_uri)


@app.route("/auth/callback")
def auth_callback():
    token = google.authorize_access_token()
    userinfo = token.get("userinfo") or google.userinfo()

    existing = db_fetchone("SELECT * FROM users WHERE google_id=?", (userinfo["sub"],))
    if not existing:
        uid = str(uuid.uuid4())
        db_execute(
            "INSERT INTO users (id,email,name,picture,google_id,access_token,refresh_token,created_at) VALUES (?,?,?,?,?,?,?,?)",
            (uid, userinfo["email"], userinfo.get("name", ""),
             userinfo.get("picture", ""), userinfo["sub"],
             token.get("access_token"), token.get("refresh_token"),
             datetime.datetime.utcnow().isoformat())
        )
        session["user_id"] = uid
    else:
        db_execute(
            "UPDATE users SET access_token=?, refresh_token=COALESCE(?,refresh_token) WHERE google_id=?",
            (token.get("access_token"), token.get("refresh_token"), userinfo["sub"])
        )
        session["user_id"] = existing["id"]
    return redirect(url_for("dashboard"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("index"))


# ---------------------------------------------------------------------------
# Routes — Dashboard
# ---------------------------------------------------------------------------
@app.route("/dashboard")
@login_required
def dashboard():
    user = current_user()
    if not user:
        return redirect(url_for("login_page"))
    invoices = db_fetchall(
        "SELECT * FROM invoices WHERE user_id=? ORDER BY processed_at DESC LIMIT 50",
        (user["id"],)
    )
    total = db_count("SELECT COUNT(*) FROM invoices WHERE user_id=?", (user["id"],))
    now = datetime.datetime.utcnow()
    month_start = datetime.datetime(now.year, now.month, 1).isoformat()
    month_count = db_count(
        "SELECT COUNT(*) FROM invoices WHERE user_id=? AND processed_at>=?",
        (user["id"], month_start)
    )
    webhook_ok = bool(
        user.get("webhook_channel_id") and user.get("webhook_expiry") and
        user["webhook_expiry"] > datetime.datetime.utcnow().isoformat()
    )
    return render_template(
        "dashboard.html", user=user, invoices=invoices,
        total=total, month_count=month_count,
        webhook_ok=webhook_ok, category_map=CATEGORY_MAP,
    )


# ---------------------------------------------------------------------------
# Routes — Drive
# ---------------------------------------------------------------------------
@app.route("/drive/folders")
@login_required
def list_drive_folders():
    user = current_user()
    service = get_drive_service(user)
    if not service:
        return jsonify({"error": "Drive not connected"}), 400
    try:
        results = service.files().list(
            q="mimeType='application/vnd.google-apps.folder' and trashed=false",
            fields="files(id,name)", pageSize=50
        ).execute()
        return jsonify({"folders": results.get("files", [])})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/drive/set-folder", methods=["POST"])
@login_required
def set_drive_folder():
    user = current_user()
    data = request.get_json()
    folder_id = data.get("folder_id")
    folder_name = data.get("folder_name", "")
    if not folder_id:
        return jsonify({"error": "folder_id required"}), 400
    db_execute(
        "UPDATE users SET drive_folder_id=?, drive_folder_name=? WHERE id=?",
        (folder_id, folder_name, user["id"])
    )
    user = current_user()
    ok = register_webhook(user)
    return jsonify({"ok": ok, "folder_name": folder_name})


@app.route("/drive/reconnect", methods=["POST"])
@login_required
def reconnect_webhook():
    user = current_user()
    if not user.get("drive_folder_id"):
        return jsonify({"error": "No folder selected"}), 400
    ok = register_webhook(user)
    return jsonify({"ok": ok})


@app.route("/webhook/<user_id>", methods=["POST"])
def receive_webhook(user_id):
    if request.headers.get("X-Goog-Resource-State", "") in ("sync", ""):
        return "", 200
    user = db_fetchone("SELECT * FROM users WHERE id=?", (user_id,))
    if not user or not user.get("drive_folder_id"):
        return "", 200
    try:
        service = get_drive_service(user)
        if not service:
            return "", 200
        processed = {r["drive_file_id"] for r in db_fetchall(
            "SELECT drive_file_id FROM invoices WHERE user_id=? AND drive_file_id IS NOT NULL",
            (user_id,)
        )}
        files = service.files().list(
            q=f"'{user['drive_folder_id']}' in parents and trashed=false",
            fields="files(id,name,mimeType)", orderBy="createdTime desc", pageSize=20
        ).execute().get("files", [])
        for f in files:
            if f["id"] in processed:
                continue
            ext = f["name"].rsplit(".", 1)[-1].lower() if "." in f["name"] else ""
            if ext not in {"pdf", "jpg", "jpeg", "png", "xlsx", "csv", "txt"}:
                continue
            try:
                req = service.files().get_media(fileId=f["id"])
                buf = io.BytesIO()
                dl = MediaIoBaseDownload(buf, req)
                done = False
                while not done:
                    _, done = dl.next_chunk()
                buf.seek(0)
                raw = buf.read()
                b64 = base64.standard_b64encode(raw).decode("utf-8")
                mime = f["mimeType"]
                blocks = []
                if mime == "application/pdf":
                    blocks.append({"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}})
                elif mime.startswith("image/"):
                    blocks.append({"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}})
                else:
                    blocks.append({"type": "text", "text": raw.decode("utf-8", errors="ignore")})
                result = run_claude(blocks)
                save_invoice(user_id, result, f["name"], drive_file_id=f["id"], source="drive")
            except Exception as e:
                app.logger.error(f"File error {f['name']}: {e}")
    except Exception as e:
        app.logger.error(f"Webhook error: {e}")
    return "", 200


# ---------------------------------------------------------------------------
# Routes — Classify + Export
# ---------------------------------------------------------------------------
@app.route("/classify", methods=["POST"])
@login_required
def classify_manual():
    user = current_user()
    content_blocks = []
    filename = "direct-text"
    try:
        if "file" in request.files and request.files["file"].filename:
            from werkzeug.utils import secure_filename
            f = request.files["file"]
            filename = secure_filename(f.filename)
            raw = f.read()
            b64 = base64.standard_b64encode(raw).decode("utf-8")
            mime = f.mimetype
            if mime == "application/pdf":
                content_blocks.append({"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}})
            elif mime.startswith("image/"):
                content_blocks.append({"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}})
            else:
                content_blocks.append({"type": "text", "text": raw.decode("utf-8", errors="ignore")})
        else:
            text = request.form.get("text", "").strip()
            if not text:
                return jsonify({"error": "Send a file or text"}), 400
            content_blocks.append({"type": "text", "text": text})
        result = run_claude(content_blocks)
        inv_id = save_invoice(user["id"], result, filename)
        return jsonify({**result, "filename": filename, "id": inv_id})
    except anthropic.AuthenticationError:
        return jsonify({"error": "Invalid Anthropic API key"}), 401
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/invoices")
@login_required
def get_invoices():
    user = current_user()
    page = int(request.args.get("page", 1))
    per_page = 20
    total = db_count("SELECT COUNT(*) FROM invoices WHERE user_id=?", (user["id"],))
    items = db_fetchall(
        "SELECT * FROM invoices WHERE user_id=? ORDER BY processed_at DESC LIMIT ? OFFSET ?",
        (user["id"], per_page, (page - 1) * per_page)
    )
    return jsonify({"total": total, "page": page, "invoices": items})


@app.route("/export")
@login_required
def export_excel():
    user = current_user()
    invoices = db_fetchall(
        "SELECT * FROM invoices WHERE user_id=? ORDER BY processed_at DESC",
        (user["id"],)
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"
    headers = ["Processed At", "File", "Category", "Vendor", "Total", "Invoice Date", "Confidence %", "Summary", "Source"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="5046E4")
    for inv in invoices:
        ws.append([
            inv.get("processed_at", ""), inv.get("filename", ""),
            CATEGORY_MAP.get(inv.get("category", ""), inv.get("category", "")),
            inv.get("vendor", ""), inv.get("total", ""), inv.get("invoice_date", ""),
            inv.get("confidence", ""), inv.get("summary", ""), inv.get("source", ""),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"invoices_{datetime.date.today()}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/health")
def health():
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
